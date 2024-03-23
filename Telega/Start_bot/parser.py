import requests
import urllib3
from bs4 import BeautifulSoup
import openpyxl
from urllib3.exceptions import InsecureRequestWarning
urllib3.disable_warnings(category=InsecureRequestWarning)
import os
import asyncio
import aiohttp

spisok = []

path = os.path.abspath("../Боты и скрипты.xlsx")
wb = openpyxl.load_workbook(path)
ws = wb.active
col = 2
for i in range(0, ws.max_row):
    for col in ws.iter_cols(1, 3):
        # print(col[i].value)
        if "https" in str(col[i].value):
            spisok.append(str(col[i].value))

for link in spisok:

    resp = requests.get(link)
    soup = BeautifulSoup(resp.text,"lxml")
    name = soup.find(class_="tgme_page_title")
    username = soup.find(class_="tgme_page_extra")
    print(name.text)
    print(username.text[4:])