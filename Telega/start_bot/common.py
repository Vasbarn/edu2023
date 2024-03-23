import requests
from bs4 import BeautifulSoup
import asyncio
import itertools
import logging
import pandas as pd
from aiogram.handlers import message
from telethon import TelegramClient, events
import socks
import os
import openpyxl

'''брать из файла excel пользователей (2 лист), отправка текстом в чат, gitignore, решение двухфакторного входа


session = "test"
api_id =  24518272
api_hash = '7d643a28aaffaab9a34a04f36da6d44c'''

session = "checkout"
api_id =  21251279
api_hash = '40aa5369202c4906ad656ad92a8d8427'
unames = []
spisok = []
alive_bots = []
id_list = []
cl = []
slovar = {"Боты":cl}
client = TelegramClient(session, api_id, api_hash,system_version="4.16.30-vxCUSTOM").start().sign_in(phone="89635769557", password="Anemon269")

chats = tuple(id_list)



@client.on(events.NewMessage(chats=chats))
async def handle_message(event):
        sender_username =  await event.get_sender()
        unames.clear()
        unames.append(sender_username.username)
        await make_file(unames)


def name_parse():
    print("it works")
    path = os.path.abspath("../Боты и скрипты.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for i in range(0, ws.max_row):
        for col in ws.iter_cols(1, 3):
            if "https" in str(col[i].value):
                spisok.append(str(col[i].value))
    for link in spisok:
        resp = requests.get(link)
        soup = BeautifulSoup(resp.text, "lxml")
        name = soup.find(class_="tgme_page_title")
        username = soup.find(class_="tgme_page_extra")
        cl.append(name.text)
        id_list.append(username.text[4:])


async def make_file(spisok: list):
        print("file_done")
        path = os.path.abspath("../Отчет о ботах.xlsx")
        if os.path.exists(path):
            os.remove(path)
        alive_bots.clear()
        for response_uname in chats:
            if response_uname in id_list:
                alive_bots.append("Работает")
            else:
                alive_bots.append("Не работает")
        slovar["Статус"] = alive_bots
        df = pd.DataFrame(slovar)
        print(df)
        df.to_excel("Отчет о ботах.xlsx", index=False)
        username = "@BiryukovaEO"
        entit = await client.get_entity(username)
        await client.send_message(entity=entit, file="Отчет о ботах.xlsx")


# async def make_file(spisok: list):
#     path = os.path.abspath("../Отчет о ботах.xlsx")
#     if os.path.exists(path):
#         os.remove(path)
#     alive_bots.clear()
#     for response_uname in chats:
#         if response_uname in id_list:
#             alive_bots.append("Работает")
#         else:
#             alive_bots.append("Не работает")
#     slovar["Статус"] = alive_bots
#     # print(slovar)
#     df = pd.DataFrame(slovar)
#     print(df)
#     df.to_excel("Отчет о ботах.xlsx", index=False)
#     username = ""
#     entit = await client.get_entity(username)
#     await client.send_message(entity=entit, file="Отчет о ботах.xlsx")


async def handle_messages_to_send():
    print("message_send")
    for elem in id_list:
        username = elem
        entit = await client.get_entity(username)
        await client.send_message(entity=entit, message= "/start")



# async def final_message_send():
#     username = "@Adrien_Order"
#     entit = await client.get_entity(username)
#     await client.send_message(entity=entit, file="ff.xlsx")

async def main():
    name_parse()
    await handle_messages_to_send()
    print("Проверка запущена")

    await client.run_until_disconnected()


if __name__ == '__main__':
    with client:

        # client.loop.run_until_complete(
        #     main()
        # )
